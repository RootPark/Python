import csv

import requests
import json
import xmltodict
import pandas as pd
from openpyxl import load_workbook


def main():
    school_code_list = []
    school_name_list = []
    df_sheet1 = pd.read_excel('seoul_school_data.xlsx', sheet_name=0,engine='openpyxl')
    df_sheet2 = pd.read_excel('seoul_school_data.xlsx', sheet_name=1,engine='openpyxl')
    school_code1 = df_sheet1.columns[3]
    school_code2 = df_sheet2.columns[3]
    school_name1 = df_sheet1.columns[4]
    school_name2 = df_sheet1.columns[4]

    for index, row in df_sheet1.iterrows():
        school_code_list.append(row[school_code1])
        school_name_list.append(row[school_name1])

    for index, row in df_sheet2.iterrows():
        school_code_list.append(row[school_code2])
        school_name_list.append(row[school_name2])

    #print(school_name_list)

    firstUrl = "https://open.neis.go.kr/hub/mealServiceDietInfo?KEY=6d3acd88db854d2d87ffe7dfb817845f&pSize=1000&ATPT_OFCDC_SC_CODE=B10&MLSV_YMD=2021&SD_SCHUL_CODE={}".format(school_code_list[0])

    content = requests.get(firstUrl).content
    dict = xmltodict.parse(content)
    data_json = json.dumps(dict['mealServiceDietInfo'])
    obj_json = json.loads(data_json)

    file = open("./dataEx.json", "w+")
    file.write(json.dumps(obj_json['row']))

    data_frame = pd.DataFrame(obj_json['row'])
    print(data_frame.count)

    df = pd.read_json("./dataEx.json")
    df.to_csv("school_food_data.csv")

    #book = load_workbook("school_food_data.xlsx")
    #writer = pd.ExcelWriter("school_food_data.xlsx", engine='openpyxl')
    #writer.book = book

    #df.to_excel(writer, sheet_name=school_name_list[0])
    #writer.save()

    for index in range(len(school_code_list)):
        with open("school_food_data.csv", 'a') as myfile:
            url = "https://open.neis.go.kr/hub/mealServiceDietInfo?KEY=6d3acd88db854d2d87ffe7dfb817845f&ATPT_OFCDC_SC_CODE=B10&MLSV_YMD=2021&pSize=1000&SD_SCHUL_CODE={}".format(school_code_list[index])

            print(index)
            content = requests.get(url).content
            dict = xmltodict.parse(content)

            try:
                data_json = json.dumps(dict['mealServiceDietInfo'])
                obj_json = json.loads(data_json)
                file = open("./dataEx.json", "w+")
                file.write(json.dumps(obj_json['row']))
                data_frame = pd.DataFrame(obj_json['row'])
                print(data_frame.count)
                df = pd.read_json("./dataEx.json")
            except:
                continue

            wt = csv.writer(myfile)
            wt.writerows(df.values.tolist())

        #book = load_workbook("school_food_data.xlsx")
        #writer = pd.ExcelWriter("school_food_data.xlsx", engine='openpyxl')
        #writer.book = book

        #df.to_excel(writer, sheet_name=school_name_list[index])
        #writer.save()

if __name__ == "__main__":
    main()