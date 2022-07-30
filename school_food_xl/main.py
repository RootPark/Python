import requests
import json
import xmltodict
import pandas as pd
from openpyxl import load_workbook
import xlsxwriter

def main():

    #schoolCodeDataToList()

    school_code_list = []
    school_name_list = []

    df_sheet1 = pd.read_excel("school_code.xlsx", sheet_name=0, engine='openpyxl')
    df_sheet2 = pd.read_excel("school_code.xlsx", sheet_name=1, engine='openpyxl')

    school_code1 = df_sheet1.columns[3]
    school_code2 = df_sheet2.columns[3]
    school_name1 = df_sheet1.columns[4]
    school_name2 = df_sheet1.columns[4]

    for index, row in df_sheet1.iterrows():
        school_code_list.append(row[school_code1])
        school_name_list.append(row[school_name1])

    for index, row in df_sheet2.iterrows():
        school_code_list.append(row[school_code2])
        school_name_list.append(row[school_name2])

    print(len(school_code_list))
    print(len(school_name_list))
    print(school_code_list)
    print(school_name_list)

    workbook = xlsxwriter.Workbook("school_food.xlsx")

    foodDataToExcel(workbook,school_code_list,school_name_list)
    workbook.close()

def schoolCodeDataToList():
    schoolCodeURL1 = "https://open.neis.go.kr/hub/schoolInfo?ATPT_OFCDC_SC_CODE=B10&KEY=6d3acd88db854d2d87ffe7dfb817845f&pSize=1000&pIndex=1"
    schoolCodeURL2 = "https://open.neis.go.kr/hub/schoolInfo?ATPT_OFCDC_SC_CODE=B10&KEY=6d3acd88db854d2d87ffe7dfb817845f&pSize=1000&pIndex=2"

    dict1 = xmltodict.parse(requests.get(schoolCodeURL1).content)
    obj1_json = json.loads(json.dumps(dict1['schoolInfo']))

    dict2 = xmltodict.parse(requests.get(schoolCodeURL2).content)
    obj2_json = json.loads(json.dumps(dict2['schoolInfo']))

    data_frame1 = pd.DataFrame(obj1_json['row'])
    data_frame2 = pd.DataFrame(obj2_json['row'])

    writer = pd.ExcelWriter("school_code.xlsx")
    data_frame1.to_excel(writer,"Sheet1")
    writer.save()

    book = load_workbook("school_code.xlsx")
    writer.book = book
    data_frame2.to_excel(writer,"Sheet2")
    writer.save()

def foodDataToExcel(workbook, school_code_list, school_name_list):
    error_count = 0
    for index in range(len(school_code_list)):
        foodURL = "https://open.neis.go.kr/hub/mealServiceDietInfo?KEY=6d3acd88db854d2d87ffe7dfb817845f&ATPT_OFCDC_SC_CODE=B10&MLSV_YMD=2021&pSize=1000&SD_SCHUL_CODE={}".format(school_code_list[index])
        school_code = school_code_list[index]
        school_name = school_name_list[index]
        print(index)
        print(school_code)
        print(school_name)


        try:
            dict = xmltodict.parse(requests.get(foodURL).content)
            obj_json = json.loads(json.dumps(dict['mealServiceDietInfo']))
            data_frame = pd.DataFrame(obj_json['row'])

            worksheet = workbook.add_worksheet(school_name_list[index])

        # 학교 코드 = school_code_list[index]
        # 학교 명 = school_name_list[index]

        # 급식 날짜 df = food_date
        # 급식 명(점심,저녁) df = food_meal
        # 급식 식수 df = food_count
        # 급식 메뉴 df = food_menu
        # 급식 칼로리 df = food_cal
        # 급식 영양정보 df = food_info
            food_date = data_frame['MLSV_YMD'].values.tolist()
            food_meal = data_frame['MMEAL_SC_NM'].values.tolist()
            food_count = data_frame['MLSV_FGR'].values.tolist()
            food_menu = data_frame['DDISH_NM'].values.tolist()
            food_cal = data_frame['CAL_INFO'].values.tolist()
            food_info = data_frame['NTR_INFO'].values.tolist()



            for idx in range(len(food_date)):
                worksheet.write(idx, 0, school_code)
                worksheet.write(idx, 1, school_name)

                worksheet.write(idx, 2, food_date[idx])
                worksheet.write(idx, 3, food_meal[idx])
                worksheet.write(idx, 4, food_count[idx])
                worksheet.write(idx, 5, food_menu[idx])
                worksheet.write(idx, 6, food_cal[idx])
                worksheet.write(idx, 7, food_info[idx])

        except:
            print("----------ERROR----------")
            error_count += 1
            continue

    print(error_count)


if __name__ == '__main__':
    main()
